from openpyxl import load_workbook, Workbook
from scipy.stats import t
import math

def filter(file_name, sheet_name, metadata_col_alphabet, control_name):
        def mean(cat, col):
                total = 0
                for row in metadata[cat]:
                        #print(table.cell(row,col).value)
                        total += float(table.cell(row,col).value)
                return total/len(metadata[cat])

        def var(cat, col):
                total = 0
                #print(table.cell(4,col).value)
                for row in metadata[cat]:
                        total += (float(table.cell(row,col).value)**2)

                return total/len(metadata[cat])-mean(cat,col)**2

        def t_test_one(Mean,SD,value):
                t_value = t.ppf(0.975,df=len(set(metadata_col))-2)
                if value>Mean+t_value*SD or value<Mean-t_value*SD:
                        return 1
                else: return 0

        def t_test_multi(Mean,SD,Mean_control,SD_control,n,n_control):
                df = (((SD**2)/n+(SD_control**2)/n_control)**2)/((((SD**2)/n)**2)/(n-1)+(((SD_control**2)/n_control)**2)/(n_control-1))
                t_value = t.ppf(0.975,df=df)
                #print("df", df, "t_value", t_value, ((Mean-Mean_control)/math.sqrt((SD**2)/n+(SD_control**2)/n_control)))
                if abs((Mean-Mean_control)/math.sqrt((SD**2)/n+(SD_control**2)/n_control))>t_value:
                        return 1
                else: return 0


        control = control_name
        file=load_workbook("../Data/"+file_name+".xlsx",data_only=True)
        table=file[sheet_name]
        write_wb=Workbook()
        write_ws=write_wb.active

        metadata_col = [i.value for i in table[metadata_col_alphabet][1:]]

        #print(metadata_control, metadata_exp);

        max_row = len(metadata_col)
        metadata = dict()

        for i in set(metadata_col):
                metadata[i]=[]

        for i in range(max_row):
                metadata[metadata_col[i]].append(i+2)

        #print(metadata)

        #labeling
        for cell in table[metadata_col_alphabet]:
                write_ws.cell(row=cell.row, column=cell.column-(ord(metadata_col_alphabet)-ord('A'))).value=cell.value

        i=3
        cur_col=3

        while table.cell(1,i).value is not None:
                flag=0
                Mean_control=mean(control,i)
                #print(var(control,i))
                SD_control=math.sqrt(var(control,i))
                for cat in (set(metadata_col)-{control}):
                        Mean=mean(cat, i)
                        SD=math.sqrt(var(cat, i))
                        if len(metadata[control])>=2 and (Mean_control!=0 or Mean!=0):
                                print(Mean_control, Mean)
                                if t_test_multi(Mean,SD,Mean_control,SD_control,len(metadata[cat]),len(metadata[control])): flag=1
                        elif len(metadata[control])==1 and (Mean_control!=0 or Mean!=0):
                                if t_test_one(Mean,SD,float(table.cell(metadata[control][0],i).value)): flag=1

                if flag!=0:
                        for col in table.iter_cols(min_row=1, max_row=max_row+1, min_col=i, max_col=i):
                                for cell in col:
                                        write_ws.cell(row=cell.row, column=cur_col-(ord(metadata_col_alphabet)-ord('A'))).value=cell.value
                        cur_col+=1
                i+=1

        write_wb.save("../Data/"+file_name+"_filtered.xlsx")
